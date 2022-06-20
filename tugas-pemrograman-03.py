## TUGAS PEMROGRAMAN 03 PENGANTAR KECERDASAN BUATAN
## KELOMPOK 3
## 1301200428 - Muhammad Rafiqi Masrur
## 1301200457 - Rasyid Riyaldi
## KELAS: IF-44-01
## Telkom University

from openpyxl import load_workbook

## Load Workbook
## Change load_workbook("File Location")
book = load_workbook("E:/Mata Kuliah/Semester 4/Pengantar Kecerdasan Buatan/Tugas/Tugas Pemrograman 03/TUGAS-PEMROGRAMAN-03-PENGANTAR-AI/traintest.xlsx")
trainSheet = book['train']
testSheet = book['test']
Phi = 3.14259
EULER = 2.71828

## Data init
class data():
    def __init__(self, id, x1, x2, x3, y):
        self.id = id    
        self.x1 = x1
        self.x2 = x2
        self.x3 = x3
        self.y = y

## Get data for train
def get_DataTrain():
    dataList = []
    i = 2
    while trainSheet["A"+str(i)].value != None:
        dataList.append(data(trainSheet["A"+str(i)].value, trainSheet["B"+str(i)].value, trainSheet["C"+str(i)].value, trainSheet["D"+str(i)].value, trainSheet["E"+str(i)].value))
        i += 1
    return dataList

## Get data for test
def get_DataTest():
    dataList = []
    i = 2
    while testSheet["A"+str(i)].value != None:
        dataList.append(data(testSheet["A"+str(i)].value, testSheet["B"+str(i)].value, testSheet["C"+str(i)].value, testSheet["D"+str(i)].value, testSheet["E"+str(i)].value))
        i += 1
    return dataList

## Calculate average value of data
def get_average(dataList):
    i = 0
    ptTrue = 0
    ptFalse = 0
    avgTrue = data(-1,0,0,0,-1)
    avgFalse = data(-1,0,0,0,-1)
    while i != len(dataList):
        if dataList[i].y == 1:
            avgTrue.x1 += dataList[i].x1
            avgTrue.x2 += dataList[i].x2
            avgTrue.x3 += dataList[i].x3
            ptTrue += 1
        else:
            avgFalse.x1 += dataList[i].x1
            avgFalse.x2 += dataList[i].x2
            avgFalse.x3 += dataList[i].x3
            ptFalse += 1
        i += 1

    avgTrue.x1 = avgTrue.x1 / ptTrue
    avgTrue.x2 = avgTrue.x2 / ptTrue
    avgTrue.x3 = avgTrue.x3 / ptTrue
    avgFalse.x1 = avgFalse.x1 / ptFalse
    avgFalse.x2 = avgFalse.x2 / ptFalse
    avgFalse.x3 = avgFalse.x3 / ptFalse

    return avgTrue, avgFalse

## Calculate standart deviation of data
def get_standDev(avgTrue, avgFalse, dataList):
    i = 0
    ptTrue = 0
    ptFalse = 0
    standDevTrue = data(-1,0,0,0,-1)
    standDevFalse = data(-1,0,0,0,-1)
    while i != len(dataList):
        if dataList[i].y == 1:
            standDevTrue.x1 = standDevTrue.x1 + ((dataList[i].x1 - avgTrue.x1) ** 2)
            standDevTrue.x2 = standDevTrue.x2 + ((dataList[i].x2 - avgTrue.x2) ** 2)
            standDevTrue.x3 = standDevTrue.x3 + ((dataList[i].x3 - avgTrue.x3) ** 2)
            ptTrue += 1
        else:
            standDevFalse.x1 = standDevFalse.x1 + ((dataList[i].x1 - avgFalse.x1) ** 2)
            standDevFalse.x2 = standDevFalse.x2 + ((dataList[i].x2 - avgFalse.x2) ** 2)
            standDevFalse.x3 = standDevFalse.x3 + ((dataList[i].x3 - avgFalse.x3) ** 2)
            ptFalse += 1
        i += 1
    
    standDevTrue.x1 = standDevTrue.x1 / ptTrue
    standDevTrue.x2 = standDevTrue.x2 / ptTrue
    standDevTrue.x3 = standDevTrue.x3 / ptTrue
    standDevTrue.x1 = standDevTrue.x1 ** 0.5
    standDevTrue.x2 = standDevTrue.x2 ** 0.5
    standDevTrue.x3 = standDevTrue.x3 ** 0.5
    standDevFalse.x1 = standDevFalse.x1 / ptFalse
    standDevFalse.x2 = standDevFalse.x2 / ptFalse
    standDevFalse.x3 = standDevFalse.x3 / ptFalse
    standDevFalse.x1 = standDevFalse.x1 ** 0.5
    standDevFalse.x2 = standDevFalse.x2 ** 0.5
    standDevFalse.x3 = standDevFalse.x3 ** 0.5

    return standDevTrue, standDevFalse

## Naive Bayes method
def naiveBayes(dataList, testData, avgTrue, avgFalse, standDevTrue, standDevFalse):
    i = 0
    sumTrue = 0
    sumFalse = 0
    while i != len(dataList):
        if dataList[i].y == 1:
            sumTrue += 1
        else:
            sumFalse += 1
        i += 1
    i = 0
    while i != len(testData):
        valTrue = (sumTrue / len(dataList)) * ((EULER ** -(((testData[i].x1 - avgTrue.x1) ** 2) / (2 * (standDevTrue.x1 ** 2)))) / (standDevTrue.x1 * ((2 * Phi) ** 0.5))) * (EULER ** -(testData[i].x2 - avgTrue.x2) ** 2) / (2 * (standDevTrue.x2 ** 2)) / (standDevTrue.x2 * ((2 * Phi) ** 0.5)) * (EULER ** -(testData[i].x3 - avgTrue.x3) ** 2) / (2 * (standDevTrue.x3 ** 2)) / (standDevTrue.x3 * ((2 * Phi) ** 0.5))
        valFalse = (sumFalse / len(dataList)) * ((EULER ** -(((testData[i].x1 - avgFalse.x1) ** 2) / (2 * (standDevFalse.x1 ** 2)))) / (standDevFalse.x1 * ((2 * Phi) ** 0.5))) * (EULER ** -(testData[i].x2 - avgFalse.x2) ** 2) / (2 * (standDevFalse.x2 ** 2)) / (standDevFalse.x2 * ((2 * Phi) ** 0.5)) * (EULER ** -(testData[i].x3 - avgFalse.x3) ** 2) / (2 * (standDevFalse.x3 ** 2)) / (standDevFalse.x3 * ((2 * Phi) ** 0.5))
        if valTrue > valFalse:
            testData[i].y = 1
        else:
            testData[i].y = 0
        i += 1
    return testData

## Import output to excel file
def outputToExcel(input):
    for i in range(len(input)):
        testSheet["E"+str(i+2)].value = input[i].y
    ## Change book.save("File Location")
    book.save("E:/Mata Kuliah/Semester 4/Pengantar Kecerdasan Buatan/Tugas/Tugas Pemrograman 03/TUGAS-PEMROGRAMAN-03-PENGANTAR-AI/traintest.xlsx")

## Training Session
print("### Training Session ###")
print("Press Any Key to start training")
input()
trainList = get_DataTrain()
avgTrue, avgFalse = get_average(trainList)
standDevTrue, standDevFalse = get_standDev(avgTrue, avgFalse, trainList)
print("### Train Complete ###")
print("")

## Testing Session
print("### Testing Session ###")
print("Press Any Key to start testing")
input()
testList = get_DataTest()
testList = naiveBayes(trainList, testList, avgTrue, avgFalse, standDevTrue, standDevFalse)
outputToExcel(testList)
print("### Test Complete ###")
print("### Data have been changed ###")
print("")